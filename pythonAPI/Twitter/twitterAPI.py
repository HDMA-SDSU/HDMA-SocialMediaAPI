class PathgeoTwitter:
    import sys
    from datetime import datetime



    '''
        createXLSX: convert tweets array into xlsx file
        input
            1. *tweets (array)
            2. *cols (array): which columns in tweets you want to export
            3. *outputPath (String)
            4. *fileName (String): with XLSX extension, such as "test.xlsx"
            5. ?keyword (string)
            6. ?sheetTitle (string)
        return filepath (string)
    '''
    def createXLSX(self, tweets, cols, outputPath, fileName, keyword=None, sheetTitle='Tweets'):
        from openpyxl import Workbook
        from BeautifulSoup import BeautifulSoup as BS
        
     
        try:
            book = Workbook()
            sheet = book.get_active_sheet()

            sheet.title = sheetTitle

            
            #create columns
            for indx, col in enumerate(cols):
                sheet.cell(row=0, column=indx).value = col.upper()

           
                #read tweets
            for rowIndx, tweet in enumerate(tweets):
                for colIndx, col in enumerate(cols):
                    if col not in tweet:
                        continue
                                
                    val = ''
                    
                    if col in ('urls', 'hashtags'):
                        if 'entities' in tweet and col in tweet['entities'] and tweet['entities'][col]:
                            if type(tweet['entities'][col][0]) in (str, unicode):
                                val = ', '.join(tweet['entities'][col])
                            elif col == 'urls':
                                val = ', '.join(map(lambda item: item['expanded_url'], tweet['entities'][col]))
                            elif col == 'hashtags':
                                val = ', '.join(map(lambda item: item['text'], tweet['entities'][col]))
                            
                    if col == 'is_retweet':
                        val = '' if 'retweeted_id' not in tweet and 'user' not in tweet else bool(tweet.get('retweeted_id', None)) 
                    
                    if col == 'retweeted_id':
                        val = tweet.get('retweeted_id', '')     
                
                    if col == 'retweet_count':
                        val = tweet.get('retweet_count', '')
                    
                    if col == 'time_zone' and 'user' in tweet:
                        val = tweet['user'].get('time_zone', '')
                    
                    if col == 'followers_count' and 'user' in tweet:
                        val = tweet['user'].get('followers_count', '')
                    
                    if col == 'friends_count' and 'user' in tweet:
                        val = tweet['user'].get('friends_count', '')
                        
                    if col == 'statuses_count' and 'user' in tweet:
                        val = tweet['user'].get('statuses_count', '')
                        
                    if col == 'language':
                        val = tweet.get('lang', None)
                        val = val or tweet.get('iso_language_code', None)
                    
                    if col == 'location':
                        if 'location' in tweet:
                            val = tweet[col]
                        elif 'user' in tweet and 'location' in tweet['user']:
                            val = tweet['user']['location']
                    
                    if col == 'from_user':
                        if 'from_user' in tweet:
                            val = tweet[col]
                        elif 'user' in tweet and type(tweet['user']) is dict and 'screen_name' in tweet['user']:
                            val = tweet['user']['screen_name']          
                    
                    if col == 'from_user_name':
                        if 'from_user_name' in tweet:
                            val = tweet[col]
                        elif 'user' in tweet and type(tweet['user']) is dict and 'name' in tweet['user']:
                            val = tweet['user']['name']
                        
                    
                    if col == "keyword":
                        val = keyword
                    elif col == "city":
                        val = ', '.join([item['name'] for item in tweet['search_info']['search_areas']])
                    elif col == "geo" and tweet['geo']:
                        val = "%f,%f" % (tweet['geo']['coordinates'][0], tweet['geo']['coordinates'][1])
                    elif col in ("created_at", "created_at_local"):
                        val = str(tweet[col])
                    elif col == 'source':
                        #strip away tags with BeautifulSoup
                        val = BS(tweet[col]).text
                    elif col in tweet:
                        val = tweet[col]

                        
                    if type(val) not in (list, dict) and col not in ("_id", 'search_info', 'entities'):
                        sheet.cell(row=rowIndx+1, column=colIndx).value = val



            book.save(outputPath+"\\"+fileName)
            return outputPath+"\\"+fileName
        
        except Exception, e:
            import traceback
            print str(e)
            print str(traceback.print_exc())







